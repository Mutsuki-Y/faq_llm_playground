"""Excel FAQ data reader module."""

import uuid
from pathlib import Path

import openpyxl
from models import Chunk, ChunkMetadata, ContentType, FAQEntry


class ExcelReader:
    """ExcelファイルからFAQデータを読み込むクラス。

    Excelのカラム構成: No., ステータス, 親カテゴリ, 子カテゴリ, タイトル, 本文
    """

    # Expected column indices (0-based)
    COL_NO = 0
    COL_STATUS = 1
    COL_PARENT_CATEGORY = 2
    COL_CHILD_CATEGORY = 3
    COL_TITLE = 4
    COL_BODY = 5

    PUBLISH_STATUS = "公開"

    def read_faq_excel(self, file_path: Path) -> list[FAQEntry]:
        """Excelファイルを読み込み、全行をFAQEntryリストとして返す。

        Args:
            file_path: Excelファイルのパス

        Returns:
            FAQEntryのリスト（ステータスに関わらず全行）

        Raises:
            FileNotFoundError: ファイルが存在しない場合
        """
        if not file_path.exists():
            raise FileNotFoundError(f"Excelファイルが見つかりません: {file_path}")

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        entries: list[FAQEntry] = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if row[self.COL_NO] is None:
                    continue

                try:
                    no_val = int(row[self.COL_NO])
                except (ValueError, TypeError):
                    no_val = row_idx - 1

                entries.append(FAQEntry(
                    no=no_val,
                    status=str(row[self.COL_STATUS] or ""),
                    parent_category=str(row[self.COL_PARENT_CATEGORY] or ""),
                    child_category=str(row[self.COL_CHILD_CATEGORY] or ""),
                    title=str(row[self.COL_TITLE] or ""),
                    body=str(row[self.COL_BODY] or ""),
                    source_file=file_path.name,
                    sheet_name=sheet_name,
                    row_number=row_idx,
                ))

        wb.close()
        return entries

    def filter_published(self, entries: list[FAQEntry]) -> list[FAQEntry]:
        """ステータスが「公開」のエントリのみをフィルタリングする。

        Args:
            entries: FAQEntryのリスト

        Returns:
            ステータスが「公開」のFAQEntryのリスト
        """
        return [e for e in entries if e.status == self.PUBLISH_STATUS]

    def faq_entry_to_chunk(self, entry: FAQEntry) -> Chunk:
        """FAQEntryをChunkに変換する（タイトル+本文結合）。

        Args:
            entry: 変換元のFAQEntry

        Returns:
            生成されたChunk
        """
        return Chunk(
            chunk_id=str(uuid.uuid4()),
            text=f"{entry.title}\n{entry.body}",
            metadata=ChunkMetadata(
                source_file=entry.source_file,
                sheet_name=entry.sheet_name,
                row_number=entry.row_number,
                parent_category=entry.parent_category,
                child_category=entry.child_category,
                title=entry.title,
                content_type=ContentType.TEXT,
            ),
        )
